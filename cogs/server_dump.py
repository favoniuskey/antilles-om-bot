import discord
from discord import app_commands
from discord.ext import commands
import json
import datetime
import io
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ServerDump(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    def _get_permission_name(self, perm: str) -> str:
        """Convertit les noms de permissions en français lisible"""
        permissions_fr = {
            "create_instant_invite": "Créer des invitations",
            "kick_members": "Expulser des membres",
            "ban_members": "Bannir des membres",
            "administrator": "Administrateur",
            "manage_channels": "Gérer les salons",
            "manage_guild": "Gérer le serveur",
            "add_reactions": "Ajouter des réactions",
            "view_audit_log": "Voir les logs d'audit",
            "priority_speaker": "Haut-parleur prioritaire",
            "stream": "Diffuser",
            "read_messages": "Lire les messages",
            "send_messages": "Envoyer des messages",
            "send_tts_messages": "Envoyer en TTS",
            "manage_messages": "Gérer les messages",
            "embed_links": "Intégrer des liens",
            "attach_files": "Joindre des fichiers",
            "read_message_history": "Voir l'historique",
            "mention_everyone": "Mentionner @everyone",
            "use_external_emojis": "Utiliser les emojis externes",
            "view_guild_insights": "Voir les statistiques",
            "connect": "Se connecter",
            "speak": "Parler",
            "mute_members": "Rendre muet",
            "deafen_members": "Assourdir",
            "move_members": "Déplacer les membres",
            "use_voice_activation": "Utiliser la VAD",
            "change_nickname": "Changer le pseudo",
            "manage_nicknames": "Gérer les pseudos",
            "manage_roles": "Gérer les rôles",
            "manage_webhooks": "Gérer les webhooks",
            "manage_emojis_and_stickers": "Gérer emojis/stickers",
            "use_application_commands": "Utiliser les commandes",
            "request_to_speak": "Demander à parler",
            "manage_events": "Gérer les événements",
            "manage_threads": "Gérer les fils",
            "create_public_threads": "Créer fils publics",
            "create_private_threads": "Créer fils privés",
            "send_messages_in_threads": "Envoyer dans fils",
            "use_external_stickers": "Utiliser stickers externes",
            "moderate_members": "Modérer les membres",
        }
        return permissions_fr.get(perm, perm)

    def _apply_header_style(self, ws, row_num):
        """Applique un style d'en-tête à une ligne"""
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        
        for cell in ws[row_num]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border

    def _apply_alternating_rows(self, ws, start_row, end_row):
        """Applique des couleurs alternées aux lignes"""
        light_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )
        
        for row_num in range(start_row, end_row + 1):
            for cell in ws[row_num]:
                if (row_num - start_row) % 2 == 0:
                    cell.fill = light_fill
                else:
                    cell.fill = white_fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _auto_adjust_columns(self, ws):
        """Ajuste automatiquement la largeur des colonnes"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    async def _generate_json_dump(self, guild: discord.Guild) -> Tuple[str, bytes]:
        """Génère le dump JSON complet du serveur"""
        server_data = {
            "server_info": {
                "id": str(guild.id),
                "name": guild.name,
                "icon_url": str(guild.icon.url) if guild.icon else None,
                "member_count": guild.member_count,
                "created_at": guild.created_at.isoformat(),
                "owner_id": str(guild.owner_id),
                "verification_level": str(guild.verification_level),
                "default_notifications": str(guild.default_notifications),
            },
            "roles": [],
            "categories": [],
            "text_channels": [],
            "voice_channels": [],
            "stage_channels": [],
            "forum_channels": [],
            "export_date": datetime.datetime.now().isoformat()
        }
        
        for role in sorted(guild.roles, key=lambda r: r.position, reverse=True):
            permissions_dict = {}
            for perm_name, perm_value in role.permissions:
                permissions_dict[perm_name] = perm_value
            
            role_data = {
                "id": str(role.id),
                "name": role.name,
                "color": role.color.value,
                "position": role.position,
                "mentionable": role.mentionable,
                "hoist": role.hoist,
                "is_default": role.is_default(),
                "is_bot_managed": role.is_bot_managed(),
                "permissions": permissions_dict,
            }
            server_data["roles"].append(role_data)
        
        for category in guild.categories:
            category_data = {
                "id": str(category.id),
                "name": category.name,
                "position": category.position,
                "permissions_overrides": []
            }
            
            for target, overwrite in category.overwrites.items():
                overwrite_data = {
                    "id": str(target.id),
                    "type": "role" if isinstance(target, discord.Role) else "member",
                    "name": target.name if isinstance(target, discord.Role) else str(target),
                    "permissions": {perm: value for perm, value in overwrite._values.items() if value is not None}
                }
                category_data["permissions_overrides"].append(overwrite_data)
            
            server_data["categories"].append(category_data)
        
        for channel in guild.text_channels:
            channel_data = {
                "id": str(channel.id),
                "name": channel.name,
                "topic": channel.topic,
                "position": channel.position,
                "category_id": str(channel.category_id) if channel.category_id else None,
                "category_name": channel.category.name if channel.category else None,
                "slowmode_delay": channel.slowmode_delay,
                "nsfw": channel.is_nsfw(),
                "permissions_overrides": []
            }
            
            for target, overwrite in channel.overwrites.items():
                overwrite_data = {
                    "id": str(target.id),
                    "type": "role" if isinstance(target, discord.Role) else "member",
                    "name": target.name if isinstance(target, discord.Role) else str(target),
                    "permissions": {perm: value for perm, value in overwrite._values.items() if value is not None}
                }
                channel_data["permissions_overrides"].append(overwrite_data)
            
            server_data["text_channels"].append(channel_data)
        
        for channel in guild.voice_channels:
            channel_data = {
                "id": str(channel.id),
                "name": channel.name,
                "position": channel.position,
                "category_id": str(channel.category_id) if channel.category_id else None,
                "category_name": channel.category.name if channel.category else None,
                "bitrate": channel.bitrate,
                "user_limit": channel.user_limit,
                "rtc_region": channel.rtc_region,
                "permissions_overrides": []
            }
            
            for target, overwrite in channel.overwrites.items():
                overwrite_data = {
                    "id": str(target.id),
                    "type": "role" if isinstance(target, discord.Role) else "member",
                    "name": target.name if isinstance(target, discord.Role) else str(target),
                    "permissions": {perm: value for perm, value in overwrite._values.items() if value is not None}
                }
                channel_data["permissions_overrides"].append(overwrite_data)
            
            server_data["voice_channels"].append(channel_data)
        
        for channel in guild.stage_channels:
            channel_data = {
                "id": str(channel.id),
                "name": channel.name,
                "position": channel.position,
                "category_id": str(channel.category_id) if channel.category_id else None,
                "category_name": channel.category.name if channel.category else None,
                "topic": channel.topic,
                "permissions_overrides": []
            }
            
            for target, overwrite in channel.overwrites.items():
                overwrite_data = {
                    "id": str(target.id),
                    "type": "role" if isinstance(target, discord.Role) else "member",
                    "name": target.name if isinstance(target, discord.Role) else str(target),
                    "permissions": {perm: value for perm, value in overwrite._values.items() if value is not None}
                }
                channel_data["permissions_overrides"].append(overwrite_data)
            
            server_data["stage_channels"].append(channel_data)
        
        for forum in guild.forums:
            forum_data = {
                "id": str(forum.id),
                "name": forum.name,
                "position": forum.position,
                "category_id": str(forum.category_id) if forum.category_id else None,
                "category_name": forum.category.name if forum.category else None,
                "permissions_overrides": []
            }
            
            for target, overwrite in forum.overwrites.items():
                overwrite_data = {
                    "id": str(target.id),
                    "type": "role" if isinstance(target, discord.Role) else "member",
                    "name": target.name if isinstance(target, discord.Role) else str(target),
                    "permissions": {perm: value for perm, value in overwrite._values.items() if value is not None}
                }
                forum_data["permissions_overrides"].append(overwrite_data)
            
            server_data["forum_channels"].append(forum_data)
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{guild.name.replace(' ', '_')}_{timestamp}_config.json"
        json_data = json.dumps(server_data, indent=2, ensure_ascii=False)
        
        return filename, json_data.encode('utf-8')

    async def _generate_excel_dump(self, guild: discord.Guild) -> Tuple[str, bytes]:
        """Génère un Excel professionnel avec plusieurs feuilles"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # ============================================================
        # FEUILLE 1: RÉSUMÉ
        # ============================================================
        ws_summary = wb.create_sheet("📊 Résumé", 0)
        
        summary_data = [
            ["RÉSUMÉ DU SERVEUR"],
            [],
            ["Nom du serveur", guild.name],
            ["ID du serveur", str(guild.id)],
            ["Propriétaire ID", str(guild.owner_id)],
            ["Nombre de membres", guild.member_count],
            [],
            ["STATISTIQUES"],
            ["Rôles", len(guild.roles)],
            ["Catégories", len(guild.categories)],
            ["Salons textuels", len(guild.text_channels)],
            ["Salons vocaux", len(guild.voice_channels)],
            ["Salons stage", len(guild.stage_channels)],
            ["Forums", len(guild.forums)],
            [],
            ["Date de création", guild.created_at.strftime("%d/%m/%Y %H:%M:%S")],
            ["Date d'export", datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
        ]
        
        for row_idx, row in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                cell.value = value
                if row_idx in [1, 8]:
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                elif col_idx == 1 and row_idx not in [1, 2, 7, 8]:
                    cell.font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        # ============================================================
        # FEUILLE 2: RÔLES
        # ============================================================
        ws_roles = wb.create_sheet("👥 Rôles", 1)
        
        headers = ['Rôle', 'ID', 'Position', 'Couleur', 'Affiché', 'Mentionable', 'Permissions']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_roles.cell(row=1, column=col_idx)
            cell.value = header
        
        self._apply_header_style(ws_roles, 1)
        ws_roles.freeze_panes = "A2"
        
        row_num = 2
        for role in sorted(guild.roles, key=lambda r: r.position, reverse=True):
            if role.is_default():
                role_name = f"{role.name} [DEFAULT]"
            elif role.is_bot_managed():
                role_name = f"{role.name} [BOT]"
            else:
                role_name = role.name
            
            perms = role.permissions
            main_perms = []
            if perms.administrator:
                main_perms.append("ADMIN")
            if perms.manage_guild:
                main_perms.append("MANAGE_SERVER")
            if perms.manage_roles:
                main_perms.append("MANAGE_ROLES")
            if perms.manage_channels:
                main_perms.append("MANAGE_CHANNELS")
            if perms.manage_messages:
                main_perms.append("MANAGE_MESSAGES")
            if perms.moderate_members:
                main_perms.append("MODERATE")
            if perms.send_messages:
                main_perms.append("SEND_MESSAGES")
            
            row_data = [
                role_name,
                str(role.id),
                str(role.position),
                f"#{role.color.value:06x}",
                "✓" if role.hoist else "✗",
                "✓" if role.mentionable else "✗",
                " | ".join(main_perms) if main_perms else "AUCUNE"
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_roles.cell(row=row_num, column=col_idx)
                cell.value = value
            
            row_num += 1
        
        self._apply_alternating_rows(ws_roles, 2, row_num - 1)
        self._auto_adjust_columns(ws_roles)
        
        # ============================================================
        # FEUILLE 3: SALONS
        # ============================================================
        ws_channels = wb.create_sheet("🏠 Salons", 2)
        
        headers = ['Catégorie', 'Salon', 'Type', 'ID', 'Position']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_channels.cell(row=1, column=col_idx)
            cell.value = header
        
        self._apply_header_style(ws_channels, 1)
        ws_channels.freeze_panes = "A2"
        
        row_num = 2
        for category in sorted(guild.categories, key=lambda c: c.position):
            # Salons textuels
            for channel in sorted([c for c in category.channels if isinstance(c, discord.TextChannel)], key=lambda c: c.position):
                row_data = [
                    category.name,
                    f"💬 {channel.name}",
                    "TEXT",
                    str(channel.id),
                    str(channel.position)
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_channels.cell(row=row_num, column=col_idx)
                    cell.value = value
                row_num += 1
            
            # Salons vocaux
            for channel in sorted([c for c in category.channels if isinstance(c, discord.VoiceChannel)], key=lambda c: c.position):
                row_data = [
                    category.name,
                    f"🔊 {channel.name}",
                    "VOICE",
                    str(channel.id),
                    str(channel.position)
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_channels.cell(row=row_num, column=col_idx)
                    cell.value = value
                row_num += 1
            
            # Salons stage
            for channel in sorted([c for c in category.channels if isinstance(c, discord.StageChannel)], key=lambda c: c.position):
                row_data = [
                    category.name,
                    f"🎤 {channel.name}",
                    "STAGE",
                    str(channel.id),
                    str(channel.position)
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_channels.cell(row=row_num, column=col_idx)
                    cell.value = value
                row_num += 1
            
            # Forums
            for channel in sorted([c for c in category.channels if isinstance(c, discord.ForumChannel)], key=lambda c: c.position):
                row_data = [
                    category.name,
                    f"📋 {channel.name}",
                    "FORUM",
                    str(channel.id),
                    str(channel.position)
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_channels.cell(row=row_num, column=col_idx)
                    cell.value = value
                row_num += 1
        
        self._apply_alternating_rows(ws_channels, 2, row_num - 1)
        self._auto_adjust_columns(ws_channels)
        
        # ============================================================
        # FEUILLE 4: PERMISSIONS (AMÉLIORÉE)
        # ============================================================
        ws_perms = wb.create_sheet("🔐 Permissions", 3)
        
        headers = ['Salon', 'Catégorie', 'Type', 'ID Salon', 'Cible', 'ID Cible', 'Type Cible', 'Permission', 'État']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_perms.cell(row=1, column=col_idx)
            cell.value = header
        
        self._apply_header_style(ws_perms, 1)
        ws_perms.freeze_panes = "A2"
        
        all_channels = (list(guild.text_channels) + list(guild.voice_channels) + 
                        list(guild.stage_channels) + list(guild.forums))
        
        row_num = 2
        for channel in all_channels:
            if channel.overwrites:
                for target, overwrite in channel.overwrites.items():
                    target_name = target.name if isinstance(target, discord.Role) else str(target)
                    target_id = str(target.id)
                    target_type = "RÔLE" if isinstance(target, discord.Role) else "MEMBRE"
                    channel_type = type(channel).__name__
                    category_name = channel.category.name if channel.category else "N/A"
                    
                    # Parcourir chaque permission explicitement
                    for perm_name, perm_value in overwrite._values.items():
                        if perm_value is not None:
                            perm_display = self._get_permission_name(perm_name)
                            state = "✅ ACCORDÉE" if perm_value else "❌ REFUSÉE"
                            
                            row_data = [
                                channel.name,
                                category_name,
                                channel_type,
                                str(channel.id),
                                target_name,
                                target_id,
                                target_type,
                                perm_display,
                                state
                            ]
                            
                            for col_idx, value in enumerate(row_data, 1):
                                cell = ws_perms.cell(row=row_num, column=col_idx)
                                cell.value = value
                                cell.alignment = Alignment(vertical="center", wrap_text=True)
                                
                                # Colorer l'état
                                if col_idx == 9:  # Colonne État
                                    if "ACCORDÉE" in str(value):
                                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                        cell.font = Font(bold=True, color="006100")
                                    else:
                                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                        cell.font = Font(bold=True, color="9C0006")
                            
                            row_num += 1
        
        self._apply_alternating_rows(ws_perms, 2, row_num - 1)
        self._auto_adjust_columns(ws_perms)
        
        # Exporter en bytes
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{guild.name.replace(' ', '_')}_{timestamp}_analysis.xlsx"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return filename, output.getvalue()

    @app_commands.command(name="dump_server_full", description="Exporte la configuration complète du serveur (JSON + Excel professionnel)")
    @app_commands.default_permissions(administrator=True)
    async def dump_server_full(self, interaction: discord.Interaction):
        """Exporte configuration JSON et Excel d'analyse pour refonte"""
        await interaction.response.defer(ephemeral=True)
        
        guild = interaction.guild
        
        try:
            # Générer les deux fichiers
            json_filename, json_data = await self._generate_json_dump(guild)
            excel_filename, excel_data = await self._generate_excel_dump(guild)
            
            # Créer les fichiers Discord
            json_file = discord.File(io.BytesIO(json_data), filename=json_filename)
            excel_file = discord.File(io.BytesIO(excel_data), filename=excel_filename)
            
            embed = discord.Embed(
                title="📊 Dump du serveur complet",
                description=f"Configuration de **{guild.name}** exportée avec succès!",
                color=discord.Color.green()
            )
            embed.add_field(name="📄 Fichiers générés:", value="✓ JSON (config complète)\n✓ Excel Pro (4 feuilles)", inline=False)
            embed.add_field(name="📈 Statistiques:", value=f"**Rôles:** {len(guild.roles)}\n**Salons:** {len(guild.text_channels) + len(guild.voice_channels)}\n**Membres:** {guild.member_count}", inline=False)
            embed.add_field(name="📋 Feuilles Excel:", value="• 📊 Résumé\n• 👥 Rôles\n• 🏠 Salons\n• 🔐 Permissions (détaillées)", inline=False)
            embed.add_field(name="💡 Utilisation:", value="- **JSON:** Restauration/backup\n- **Excel:** Analyse visuelle pour refonte", inline=False)
            embed.set_footer(text="À partager avec les admins pour la phase théorique de refonte")
            
            await interaction.followup.send(embed=embed, files=[json_file, excel_file], ephemeral=True)
            
        except Exception as e:
            await interaction.followup.send(f"❌ Erreur lors de l'export: {str(e)}", ephemeral=True)

async def setup(bot):
    await bot.add_cog(ServerDump(bot))